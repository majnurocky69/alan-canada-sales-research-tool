import pandas as pd
import os
import datetime
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

EXCEL_FILE = "Master_Rankings.xlsx"
EXPORTS_DIR = "exports"

def ensure_export_dir():
    if not os.path.exists(EXPORTS_DIR):
        os.makedirs(EXPORTS_DIR)

def inject_visual_analytics(filepath):
    """
    Opens the newly minted Pandas Excel file internally, clears old chart layers,
    and dynamically draws a strictly formatted Bar Chart projecting the Propensity scores natively into the corner.
    """
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Strip old visual data layers out to avoid messy overlapping/crashing
        if hasattr(ws, '_charts'):
            ws._charts = []
            
        max_row = ws.max_row
        if max_row < 2: 
            return # Only a header exists, no mathematical comparisons to draw yet
             
        chart = BarChart()
        chart.type = "col"
        chart.style = 13 # Sharp corporate blue style
        chart.title = "Target Lead Propensity (Alan ICP Match)"
        chart.y_axis.title = "Score / 100"
        chart.x_axis.title = "Target Company"
        chart.legend = None # Remove generic unhelpful legend tags
        
        # Map X-Axis Labels (Column B: Company_Name) and Y-Axis Weights (Column D: Propensity_Score)
        # Note: Excel expects 1-indexed integers for Reference classes
        score_data = Reference(ws, min_col=4, min_row=1, max_row=max_row, max_col=4)
        company_categories = Reference(ws, min_col=2, min_row=2, max_row=max_row)
        
        chart.add_data(score_data, titles_from_data=True)
        chart.set_categories(company_categories)
        
        chart.width = 25 # Emphasize it visually wide across the secondary spreadsheet region
        chart.height = 12
        
        # Throw the visual dashboard natively onto Column O, Row 2 (Right of the text tables)
        ws.add_chart(chart, "O2")
        
        wb.save(filepath)
    except Exception as e:
        print(f"Visual Analytical Execution Alert: {str(e)}")


def export_to_excel(company_name, apollo_data, ai_evaluation):
    """
    Flattens the dense dictionary blocks into a static Pandas DataFrame row.
    Automatically checks if Master_Rankings.xlsx exists.
    If yes, appends internally. If no, creates new structure.
    After the file saves to disk, it triggers `inject_visual_analytics(EXCEL_FILE)`.
    """
    row_data = {
        "Scanned_Date": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "Company_Name": company_name,
        "Domain": apollo_data.get("domain", "Unknown"),
        "Propensity_Score": ai_evaluation.get("total_score", "Error"),
        "Employee_Count": apollo_data.get("employee_count", "Unknown"),
        "Location": apollo_data.get("location", "Unknown"),
        "Industry": apollo_data.get("industry", "Unknown"),
        "Annual_Revenue": apollo_data.get("annual_revenue", "Unknown"),
        "Total_Funding": apollo_data.get("total_funding", "Unknown"),
        "Executive_Summary": ai_evaluation.get("executive_summary", ""),
        "Historical_Trend": ai_evaluation.get("historical_trend", ""),
        "Score_Justification": ai_evaluation.get("score_justification", "")
    }
    
    df = pd.DataFrame([row_data])
    
    try:
        if os.path.exists(EXCEL_FILE):
             existing_df = pd.read_excel(EXCEL_FILE)
             
             # Natively convert incoming Score to numeric to avoid chart data-type crashes
             df['Propensity_Score'] = pd.to_numeric(df['Propensity_Score'], errors='coerce')
             
             updated_df = pd.concat([existing_df, df], ignore_index=True)
             
             # Natively convert previous scores to numeric as well just in case they were strings visually
             updated_df['Propensity_Score'] = pd.to_numeric(updated_df['Propensity_Score'], errors='coerce')
             
             # Clean out deduplication to avoid redundant UI chart bars
             updated_df = updated_df.drop_duplicates(subset=['Company_Name', 'Scanned_Date'], keep='last')
             
             updated_df.to_excel(EXCEL_FILE, index=False)
        else:
             df['Propensity_Score'] = pd.to_numeric(df['Propensity_Score'], errors='coerce')
             df.to_excel(EXCEL_FILE, index=False)
             
        # Inject the automated embedded chart layout into the newly saved Pandas frame
        inject_visual_analytics(EXCEL_FILE)
             
        return {"success": True, "filepath": EXCEL_FILE}
    except Exception as e:
        return {"success": False, "error": str(e)}

def export_to_markdown(company_name, ai_evaluation):
    """
    Writes out a completely raw Claude/ChatGPT formatted Markdown document containing all 
    citations, contacts, and custom talking points strictly for SDR automation workflows.
    """
    ensure_export_dir()
    
    safe_name = company_name.replace(" ", "_").replace("/", "").replace("\\", "").replace(".", "")
    filename = f"{EXPORTS_DIR}/{safe_name}_Claude_Context.md"
    
    content = f"# Alan Canada Sales Context: {company_name}\n\n"
    content += f"**Target Propensity Score:** {ai_evaluation.get('total_score', 'Unknown')}/100\n\n"
    
    content += "## 1. Executive Summary & Growth Trajectory\n"
    content += f"{ai_evaluation.get('executive_summary', 'None provided.')}\n\n"
    
    content += "## 2. Artificial Intelligence Strategic Angles\n"
    content += "> *Note: Use these angles to construct your cold outreach sequences.*\n\n"
    
    talking_points = ai_evaluation.get('talking_points', [])
    if talking_points:
        for tp in talking_points:
            content += f"### {tp.get('title', 'Point')}\n{tp.get('content', '')}\n\n"
    else:
        content += "No actionable talking points generated.\n\n"
    
    content += "## 3. Referenced Deep Web Citations\n"
    citations = ai_evaluation.get('citations', [])
    if citations:
         for c in citations:
              content += f"- **{c.get('source', 'Source')}**: {c.get('insight_extracted', '')} ([Link]({c.get('url', '#')}))\n"
    else:
         content += "No deep citations established.\n"
         
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(content)
        return {"success": True, "filepath": filename}
    except Exception as e:
        return {"success": False, "error": str(e)}
